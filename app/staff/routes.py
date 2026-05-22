"""
Staff Blueprint — IT Staff can only see their own assigned tickets.
"""

import io
from datetime import datetime, timezone
from functools import wraps

from flask import (Blueprint, render_template, redirect,
                   url_for, flash, abort, request, send_file)
from flask_login import login_required, current_user

from app.extensions import db
from app.models import Ticket, TicketStatus, UserRole
from app.staff.forms import ResolveTicketForm, StartProgressForm
from app.email_service.notifications import send_ticket_solved_email

staff_bp = Blueprint('staff', __name__, template_folder='../templates/staff')


def staff_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            abort(403)
        if current_user.role not in (UserRole.STAFF, UserRole.ADMIN):
            abort(403)
        return f(*args, **kwargs)
    return decorated_function


@staff_bp.route('/')
@login_required
@staff_required
def dashboard():
    if current_user.is_admin:
        open_tickets = Ticket.query.filter(
            Ticket.status != TicketStatus.SOLVED,
            Ticket.status != TicketStatus.CLOSED,
        ).order_by(Ticket.submitted_at.desc()).all()
        solved_tickets = Ticket.query.filter(
            Ticket.solved_by_id == current_user.id
        ).order_by(Ticket.solved_at.desc()).limit(5).all()
    else:
        open_tickets = Ticket.query.filter_by(
            assigned_to_id=current_user.id
        ).filter(
            Ticket.status != TicketStatus.SOLVED,
            Ticket.status != TicketStatus.CLOSED,
        ).order_by(Ticket.submitted_at.desc()).all()

        solved_tickets = Ticket.query.filter_by(
            assigned_to_id=current_user.id,
            status=TicketStatus.SOLVED,
        ).order_by(Ticket.solved_at.desc()).limit(5).all()

    # Stats for this staff member
    stats = {
        'total_assigned': Ticket.query.filter_by(assigned_to_id=current_user.id).count(),
        'open':           Ticket.query.filter_by(assigned_to_id=current_user.id).filter(
                            Ticket.status == TicketStatus.OPEN).count(),
        'in_progress':    Ticket.query.filter_by(assigned_to_id=current_user.id).filter(
                            Ticket.status == TicketStatus.IN_PROGRESS).count(),
        'solved':         Ticket.query.filter_by(solved_by_id=current_user.id).filter(
                            Ticket.status == TicketStatus.SOLVED).count(),
    }

    return render_template(
        'staff/dashboard.html',
        open_tickets=open_tickets,
        solved_tickets=solved_tickets,
        stats=stats,
    )


@staff_bp.route('/ticket/<int:ticket_id>')
@login_required
@staff_required
def ticket_detail(ticket_id: int):
    ticket = db.session.get(Ticket, ticket_id) or abort(404)

    if not current_user.is_admin:
        if ticket.assigned_to_id != current_user.id:
            abort(403)

    resolve_form   = ResolveTicketForm()
    progress_form  = StartProgressForm()
    return render_template('staff/ticket_detail.html', ticket=ticket,
                           form=resolve_form, progress_form=progress_form)


@staff_bp.route('/ticket/<int:ticket_id>/progress', methods=['POST'])
@login_required
@staff_required
def mark_in_progress(ticket_id: int):
    ticket = db.session.get(Ticket, ticket_id) or abort(404)

    if not current_user.is_admin and ticket.assigned_to_id != current_user.id:
        abort(403)

    form = StartProgressForm()

    if ticket.status != TicketStatus.ASSIGNED:
        flash('Cannot change status to In Progress from current state.', 'warning')
        return redirect(url_for('staff.ticket_detail', ticket_id=ticket_id))

    if form.validate_on_submit():
        form.sanitize()
        ticket.status                    = TicketStatus.IN_PROGRESS
        ticket.in_progress_at            = datetime.now(timezone.utc)
        ticket.estimated_resolution_time = form.estimated_resolution_time.data
        db.session.commit()
        flash(f'Ticket {ticket.ticket_ref} is now In Progress.', 'info')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(error, 'danger')

    return redirect(url_for('staff.ticket_detail', ticket_id=ticket_id))


@staff_bp.route('/ticket/<int:ticket_id>/solve', methods=['POST'])
@login_required
@staff_required
def solve_ticket(ticket_id: int):
    ticket = db.session.get(Ticket, ticket_id) or abort(404)

    if not current_user.is_admin and ticket.assigned_to_id != current_user.id:
        abort(403)

    form = ResolveTicketForm()

    if form.validate_on_submit():
        form.sanitize()
        try:
            ticket.mark_solved(resolved_by_user=current_user,
                               remark=form.resolution_remark.data)
            db.session.commit()
            send_ticket_solved_email(ticket)
            flash(f'Ticket {ticket.ticket_ref} marked as Solved.', 'success')
        except ValueError as e:
            flash(str(e), 'danger')
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(error, 'danger')

    return redirect(url_for('staff.ticket_detail', ticket_id=ticket_id))


# ---------------------------------------------------------------------------
# Solved Tickets History — Staff sees only their own resolved tickets
# ---------------------------------------------------------------------------

@staff_bp.route('/history')
@login_required
@staff_required
def solved_history():
    date_from_str = request.args.get('date_from', '')
    date_to_str   = request.args.get('date_to',   '')
    page          = request.args.get('page', 1, type=int)

    query = Ticket.query.filter(
        Ticket.solved_by_id == current_user.id,
        Ticket.status       == TicketStatus.SOLVED,
    )

    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').replace(tzinfo=timezone.utc)
            query = query.filter(Ticket.solved_at >= date_from)
        except ValueError:
            pass

    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').replace(
                hour=23, minute=59, second=59, tzinfo=timezone.utc)
            query = query.filter(Ticket.solved_at <= date_to)
        except ValueError:
            pass

    query   = query.order_by(Ticket.solved_at.desc())
    tickets = query.paginate(page=page, per_page=25, error_out=False)

    return render_template(
        'staff/solved_history.html',
        tickets=tickets,
        date_from=date_from_str,
        date_to=date_to_str,
        total_count=tickets.total,
    )


@staff_bp.route('/history/export')
@login_required
@staff_required
def export_solved_history():
    """Export this staff member's solved tickets to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    date_from_str = request.args.get('date_from', '')
    date_to_str   = request.args.get('date_to',   '')

    query = Ticket.query.filter(
        Ticket.solved_by_id == current_user.id,
        Ticket.status       == TicketStatus.SOLVED,
    )

    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').replace(tzinfo=timezone.utc)
            query = query.filter(Ticket.solved_at >= date_from)
        except ValueError:
            pass

    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').replace(
                hour=23, minute=59, second=59, tzinfo=timezone.utc)
            query = query.filter(Ticket.solved_at <= date_to)
        except ValueError:
            pass

    tickets = query.order_by(Ticket.solved_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'My Solved Tickets'

    header_fill = PatternFill(start_color='1a3c5e', end_color='1a3c5e', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    headers = [
        'Ticket Ref', 'Submitted By', 'Employee ID', 'Department',
        'Category', 'Sub-Category', 'Asset ID', 'Priority', 'Problem Description',
        'Assigned At', 'Resolved On', 'Resolution Time', 'Resolution Remark'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center')

    def _fmt_resolution_time(t):
        if t.solved_at and t.assigned_at:
            total_sec = int((t.solved_at - t.assigned_at).total_seconds())
            if total_sec < 0:
                return 'N/A'
            d = total_sec // 86400
            h = (total_sec % 86400) // 3600
            m = (total_sec % 3600) // 60
            parts = []
            if d: parts.append(f'{d}d')
            if h: parts.append(f'{h}h')
            parts.append(f'{m}m')
            return ' '.join(parts)
        return 'N/A'

    for row, ticket in enumerate(tickets, 2):
        ws.cell(row=row, column=1,  value=ticket.ticket_ref)
        ws.cell(row=row, column=2,  value=ticket.submitter_name)
        ws.cell(row=row, column=3,  value=ticket.submitter_employee_id)
        ws.cell(row=row, column=4,  value=ticket.submitter_department)
        ws.cell(row=row, column=5,  value=ticket.category.value if ticket.category else '')
        ws.cell(row=row, column=6,  value=ticket.sub_category or '')
        ws.cell(row=row, column=7,  value=ticket.asset_id or '')
        ws.cell(row=row, column=8,  value=ticket.priority.value)
        ws.cell(row=row, column=9,  value=ticket.problem_description)
        ws.cell(row=row, column=10, value=ticket.assigned_at.strftime('%d %b %Y %H:%M') if ticket.assigned_at else '')
        ws.cell(row=row, column=11, value=ticket.solved_at.strftime('%d %b %Y %H:%M') if ticket.solved_at else '')
        ws.cell(row=row, column=12, value=_fmt_resolution_time(ticket))
        ws.cell(row=row, column=13, value=ticket.resolution_remark or '')

    for col in ws.columns:
        max_length = max((len(str(cell.value or '')) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"my_solved_tickets_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
