"""
Admin Blueprint — fully protected, Admin role only.
"""

import io
from datetime import datetime, timezone
from functools import wraps

from flask import (Blueprint, render_template, redirect,
                   url_for, flash, request, abort, send_file)
from flask_login import login_required, current_user

from app.extensions import db
from app.models import User, Ticket, TicketStatus, TicketPriority, TicketCategory, UserRole
from app.admin.forms import (CreateUserForm, EditUserForm, AssignTicketForm, ChangePriorityForm,
                              SolveTicketForm, AdminResetPasswordForm)
from app.email_service.notifications import send_ticket_solved_email

admin_bp = Blueprint('admin', __name__, template_folder='../templates/admin')


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            abort(403)
        return f(*args, **kwargs)
    return decorated_function


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

@admin_bp.route('/')
@login_required
@admin_required
def dashboard():
    status_filter   = request.args.get('status',   '')
    priority_filter = request.args.get('priority', '')
    category_filter = request.args.get('category', '')
    sort_by         = request.args.get('sort',     'submitted_at')
    order           = request.args.get('order',    'desc')
    page            = request.args.get('page', 1, type=int)

    query = Ticket.query

    if status_filter:
        try:
            query = query.filter(Ticket.status == TicketStatus(status_filter))
        except ValueError:
            pass

    if priority_filter:
        try:
            query = query.filter(Ticket.priority == TicketPriority(priority_filter))
        except ValueError:
            pass

    if category_filter:
        try:
            query = query.filter(Ticket.category == TicketCategory(category_filter))
        except ValueError:
            pass

    allowed_sort_columns = {
        'submitted_at': Ticket.submitted_at,
        'priority':     Ticket.priority,
        'status':       Ticket.status,
    }
    sort_column = allowed_sort_columns.get(sort_by, Ticket.submitted_at)
    query = query.order_by(sort_column.asc() if order == 'asc' else sort_column.desc())

    tickets = query.paginate(page=page, per_page=20, error_out=False)

    stats = {
        'total':       Ticket.query.count(),
        'open':        Ticket.query.filter_by(status=TicketStatus.OPEN).count(),
        'assigned':    Ticket.query.filter_by(status=TicketStatus.ASSIGNED).count(),
        'in_progress': Ticket.query.filter_by(status=TicketStatus.IN_PROGRESS).count(),
        'solved':      Ticket.query.filter_by(status=TicketStatus.SOLVED).count(),
        'high':        Ticket.query.filter_by(priority=TicketPriority.HIGH).count(),
    }

    category_stats = {
        'Hardware':    Ticket.query.filter_by(category=TicketCategory.HARDWARE).count(),
        'Software':    Ticket.query.filter_by(category=TicketCategory.SOFTWARE).count(),
        'Network':     Ticket.query.filter_by(category=TicketCategory.NETWORK).count(),
        'CCTV':        Ticket.query.filter_by(category=TicketCategory.CCTV).count(),
        'Application': Ticket.query.filter_by(category=TicketCategory.APPLICATION).count(),
    }

    return render_template(
        'admin/dashboard.html',
        tickets=tickets,
        stats=stats,
        category_stats=category_stats,
        status_filter=status_filter,
        priority_filter=priority_filter,
        category_filter=category_filter,
    )


# ---------------------------------------------------------------------------
# Ticket Detail
# ---------------------------------------------------------------------------

@admin_bp.route('/ticket/<int:ticket_id>', methods=['GET'])
@login_required
@admin_required
def ticket_detail(ticket_id: int):
    ticket        = db.session.get(Ticket, ticket_id) or abort(404)
    staff_members = User.query.filter(
        User.role.in_([UserRole.STAFF, UserRole.ADMIN]),
        User.is_active == True
    ).all()

    assign_form   = AssignTicketForm()
    assign_form.assigned_to.choices = [
        (s.id, f'{s.full_name} ({s.employee_id})') for s in staff_members
    ]
    priority_form = ChangePriorityForm(priority=ticket.priority.value)
    solve_form    = SolveTicketForm()

    return render_template(
        'admin/ticket_detail.html',
        ticket=ticket,
        assign_form=assign_form,
        priority_form=priority_form,
        solve_form=solve_form,
    )


@admin_bp.route('/ticket/<int:ticket_id>/assign', methods=['POST'])
@login_required
@admin_required
def assign_ticket(ticket_id: int):
    ticket        = db.session.get(Ticket, ticket_id) or abort(404)
    staff_members = User.query.filter(
        User.role.in_([UserRole.STAFF, UserRole.ADMIN]),
        User.is_active == True
    ).all()

    form = AssignTicketForm()
    form.assigned_to.choices = [
        (s.id, f'{s.full_name} ({s.employee_id})') for s in staff_members
    ]

    # Block reassignment if In Progress or beyond
    locked_statuses = [TicketStatus.IN_PROGRESS, TicketStatus.SOLVED, TicketStatus.CLOSED]
    if ticket.status in locked_statuses:
        flash(f'Cannot reassign — ticket is already {ticket.status.value}.', 'warning')
        return redirect(url_for('admin.ticket_detail', ticket_id=ticket_id))

    if form.validate_on_submit():
        staff_user = db.session.get(User, form.assigned_to.data)
        if not staff_user:
            flash('Selected staff member not found.', 'danger')
            return redirect(url_for('admin.ticket_detail', ticket_id=ticket_id))
        ticket.assign_to(staff_user)
        db.session.commit()
        flash(f'Ticket {ticket.ticket_ref} assigned to {staff_user.full_name}.', 'success')
    else:
        flash('Assignment failed.', 'danger')

    return redirect(url_for('admin.ticket_detail', ticket_id=ticket_id))


@admin_bp.route('/ticket/<int:ticket_id>/priority', methods=['POST'])
@login_required
@admin_required
def change_priority(ticket_id: int):
    ticket = db.session.get(Ticket, ticket_id) or abort(404)
    form   = ChangePriorityForm()

    if form.validate_on_submit():
        priority_map = {
            'Low':    TicketPriority.LOW,
            'Medium': TicketPriority.MEDIUM,
            'High':   TicketPriority.HIGH,
        }
        new_priority = priority_map.get(form.priority.data)
        if new_priority:
            ticket.priority = new_priority
            db.session.commit()
            flash(f'Priority updated to {new_priority.value}.', 'success')

    return redirect(url_for('admin.ticket_detail', ticket_id=ticket_id))


@admin_bp.route('/ticket/<int:ticket_id>/solve', methods=['POST'])
@login_required
@admin_required
def solve_ticket(ticket_id: int):
    ticket = db.session.get(Ticket, ticket_id) or abort(404)
    form   = SolveTicketForm()

    if form.validate_on_submit():
        form.sanitize()
        try:
            ticket.mark_solved(resolved_by_user=current_user,
                               remark=form.resolution_remark.data)
            db.session.commit()
            send_ticket_solved_email(ticket)
            flash(f'Ticket {ticket.ticket_ref} marked as Solved.', 'success')
        except ValueError as e:
            flash(str(e), 'danger')

    return redirect(url_for('admin.ticket_detail', ticket_id=ticket_id))


# ---------------------------------------------------------------------------
# User Management
# ---------------------------------------------------------------------------

@admin_bp.route('/users')
@login_required
@admin_required
def list_users():
    role_filter = request.args.get('role', '')
    query       = User.query

    if role_filter:
        try:
            query = query.filter_by(role=UserRole(role_filter))
        except ValueError:
            pass

    users = query.order_by(User.created_at.desc()).all()
    return render_template('admin/users.html', users=users, role_filter=role_filter)


@admin_bp.route('/users/create', methods=['GET', 'POST'])
@login_required
@admin_required
def create_user():
    form = CreateUserForm()

    if form.validate_on_submit():
        form.sanitize()

        if User.query.filter_by(email=form.email.data).first():
            flash('An account with this email already exists.', 'danger')
            return render_template('admin/create_user.html', form=form)

        if User.query.filter_by(employee_id=form.employee_id.data).first():
            flash('An account with this Employee ID already exists.', 'danger')
            return render_template('admin/create_user.html', form=form)

        role_map = {
            'employee': UserRole.EMPLOYEE,
            'staff':    UserRole.STAFF,
            'admin':    UserRole.ADMIN,
        }
        role = role_map.get(form.role.data, UserRole.EMPLOYEE)

        new_user = User(
            full_name            = form.full_name.data,
            employee_id          = form.employee_id.data,
            email                = form.email.data,
            department           = form.department.data,
            role                 = role,
            must_change_password = True,  # Force password change on first login
        )
        new_user.password = form.password.data
        db.session.add(new_user)
        db.session.commit()

        flash(
            f'Account created for {new_user.full_name} ({role.value}). '
            f'They will be asked to change their password on first login.',
            'success'
        )
        return redirect(url_for('admin.list_users'))

    return render_template('admin/create_user.html', form=form)


@admin_bp.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(user_id: int):
    user = db.session.get(User, user_id) or abort(404)
    form = EditUserForm(obj=user)

    # Pre-populate role field from enum value
    if request.method == 'GET':
        form.role.data = user.role.value

    if form.validate_on_submit():
        form.sanitize()

        # Check email uniqueness (exclude current user)
        existing_email = User.query.filter(
            User.email == form.email.data, User.id != user.id
        ).first()
        if existing_email:
            flash('Another account with this email already exists.', 'danger')
            return render_template('admin/edit_user.html', form=form, user=user)

        # Check employee_id uniqueness (exclude current user)
        existing_eid = User.query.filter(
            User.employee_id == form.employee_id.data, User.id != user.id
        ).first()
        if existing_eid:
            flash('Another account with this Employee ID already exists.', 'danger')
            return render_template('admin/edit_user.html', form=form, user=user)

        role_map = {
            'employee': UserRole.EMPLOYEE,
            'staff':    UserRole.STAFF,
            'admin':    UserRole.ADMIN,
        }
        user.full_name   = form.full_name.data
        user.employee_id = form.employee_id.data
        user.email       = form.email.data
        user.department  = form.department.data
        user.role        = role_map.get(form.role.data, UserRole.EMPLOYEE)
        db.session.commit()

        flash(f'Account for {user.full_name} updated successfully.', 'success')
        return redirect(url_for('admin.list_users'))

    return render_template('admin/edit_user.html', form=form, user=user)


@admin_bp.route('/users/<int:user_id>/toggle', methods=['POST'])
@login_required
@admin_required
def toggle_user(user_id: int):
    user = db.session.get(User, user_id) or abort(404)

    if user.id == current_user.id:
        flash('You cannot deactivate your own account.', 'warning')
        return redirect(url_for('admin.list_users'))

    user.is_active = not user.is_active
    db.session.commit()
    state = 'activated' if user.is_active else 'deactivated'
    flash(f'Account for {user.full_name} has been {state}.', 'info')
    return redirect(url_for('admin.list_users'))


@admin_bp.route('/users/<int:user_id>/reset-password', methods=['GET', 'POST'])
@login_required
@admin_required
def reset_user_password(user_id: int):
    """Admin resets a user's password — forces them to change it on next login."""
    user = db.session.get(User, user_id) or abort(404)
    form = AdminResetPasswordForm()

    if form.validate_on_submit():
        user.password             = form.new_password.data
        user.must_change_password = True
        db.session.commit()
        flash(
            f'Password reset for {user.full_name}. '
            f'They will be asked to change it on next login.',
            'success'
        )
        return redirect(url_for('admin.list_users'))

    return render_template('admin/reset_password.html', form=form, user=user)


@admin_bp.route('/users/<int:user_id>/tickets')
@login_required
@admin_required
def user_tickets(user_id: int):
    """Admin views all tickets ever submitted by a specific employee."""
    user    = db.session.get(User, user_id) or abort(404)
    tickets = Ticket.query.filter_by(
        created_by_id=user_id
    ).order_by(Ticket.submitted_at.desc()).all()

    return render_template('admin/user_tickets.html', user=user, tickets=tickets)


# ---------------------------------------------------------------------------
# Solved Tickets History
# ---------------------------------------------------------------------------

@admin_bp.route('/history')
@login_required
@admin_required
def solved_history():
    staff_filter  = request.args.get('staff_id', '', type=str)
    date_from_str = request.args.get('date_from', '')
    date_to_str   = request.args.get('date_to', '')
    page          = request.args.get('page', 1, type=int)

    query = Ticket.query.filter(Ticket.status == TicketStatus.SOLVED)

    if staff_filter:
        try:
            query = query.filter(Ticket.solved_by_id == int(staff_filter))
        except ValueError:
            pass

    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').replace(tzinfo=timezone.utc)
            query = query.filter(Ticket.solved_at >= date_from)
        except ValueError:
            pass

    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').replace(
                hour=23, minute=59, second=59, tzinfo=timezone.utc)
            query = query.filter(Ticket.solved_at <= date_to)
        except ValueError:
            pass

    query   = query.order_by(Ticket.solved_at.desc())
    tickets = query.paginate(page=page, per_page=25, error_out=False)

    staff_members = User.query.filter(
        User.role.in_([UserRole.STAFF, UserRole.ADMIN]),
        User.is_active == True
    ).order_by(User.full_name).all()

    return render_template(
        'admin/solved_history.html',
        tickets=tickets,
        staff_members=staff_members,
        staff_filter=staff_filter,
        date_from=date_from_str,
        date_to=date_to_str,
        total_count=tickets.total,
    )


@admin_bp.route('/history/export')
@login_required
@admin_required
def export_solved_history():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    staff_filter  = request.args.get('staff_id', '')
    date_from_str = request.args.get('date_from', '')
    date_to_str   = request.args.get('date_to', '')

    query = Ticket.query.filter(Ticket.status == TicketStatus.SOLVED)

    if staff_filter:
        try:
            query = query.filter(Ticket.solved_by_id == int(staff_filter))
        except ValueError:
            pass

    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').replace(tzinfo=timezone.utc)
            query = query.filter(Ticket.solved_at >= date_from)
        except ValueError:
            pass

    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').replace(
                hour=23, minute=59, second=59, tzinfo=timezone.utc)
            query = query.filter(Ticket.solved_at <= date_to)
        except ValueError:
            pass

    tickets = query.order_by(Ticket.solved_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Solved Tickets'

    header_fill = PatternFill(start_color='1a3c5e', end_color='1a3c5e', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    headers = [
        'Ticket Ref', 'Raised By', 'Employee ID', 'Department',
        'Category', 'Sub-Category', 'Asset ID', 'Priority', 'Problem Description',
        'Assigned At', 'Resolved By', 'Resolved On', 'Resolution Time',
        'Resolution Remark'
    ]

    from openpyxl.styles import Alignment
    for col, header in enumerate(headers, 1):
        cell            = ws.cell(row=1, column=col, value=header)
        cell.fill       = header_fill
        cell.font       = header_font
        cell.alignment  = Alignment(horizontal='center')

    def _fmt_resolution_time(t):
        if t.solved_at and t.assigned_at:
            total_sec = int((t.solved_at - t.assigned_at).total_seconds())
            if total_sec < 0:
                return 'N/A'
            d = total_sec // 86400
            h = (total_sec % 86400) // 3600
            m = (total_sec % 3600) // 60
            parts = []
            if d: parts.append(f'{d}d')
            if h: parts.append(f'{h}h')
            parts.append(f'{m}m')
            return ' '.join(parts)
        return 'N/A'

    for row, ticket in enumerate(tickets, 2):
        ws.cell(row=row, column=1,  value=ticket.ticket_ref)
        ws.cell(row=row, column=2,  value=ticket.submitter_name)
        ws.cell(row=row, column=3,  value=ticket.submitter_employee_id)
        ws.cell(row=row, column=4,  value=ticket.submitter_department)
        ws.cell(row=row, column=5,  value=ticket.category.value if ticket.category else '')
        ws.cell(row=row, column=6,  value=ticket.sub_category or '')
        ws.cell(row=row, column=7,  value=ticket.asset_id or '')
        ws.cell(row=row, column=8,  value=ticket.priority.value)
        ws.cell(row=row, column=9,  value=ticket.problem_description)
        ws.cell(row=row, column=10, value=ticket.assigned_at.strftime('%d %b %Y %H:%M') if ticket.assigned_at else '')
        ws.cell(row=row, column=11, value=ticket.solved_by.full_name if ticket.solved_by else '')
        ws.cell(row=row, column=12, value=ticket.solved_at.strftime('%d %b %Y %H:%M') if ticket.solved_at else '')
        ws.cell(row=row, column=13, value=_fmt_resolution_time(ticket))
        ws.cell(row=row, column=14, value=ticket.resolution_remark or '')

    for col in ws.columns:
        max_length = max((len(str(cell.value or '')) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"solved_tickets_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buffer, as_attachment=True, download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
